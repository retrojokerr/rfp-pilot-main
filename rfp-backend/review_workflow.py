"""
review_workflow.py — Phase 1 of the document-centric review workflow.

A submission is a snapshot of a sheet + its answers, handed from a submitter
to a reviewer as one unit. See REVIEW_WORKFLOW_DESIGN.md.

KB ingestion invariant:
  - corrected items ingest to Qdrant ONLY here, on approval
  - rejected items never ingest
  - "accepted" (path-1, KB-direct) items already ingested at generation time
"""
from __future__ import annotations
from typing import Optional
from datetime import datetime, timezone

from fastapi import APIRouter, Body, Depends, HTTPException
from pydantic import BaseModel, Field
from sqlmodel import Session, select

from io import BytesIO
from pathlib import Path
import re

from fastapi.responses import StreamingResponse

from database import engine
from models import ReviewSubmission, ReviewItem, Notification, OriginalDocument, FeedbackPair
from auth import require, current_user, User

router = APIRouter(prefix="/review", tags=["review"])


def _now():
    return datetime.now(timezone.utc)


class ReviewItemIn(BaseModel):
    question_id: str
    question: str
    section: Optional[str] = None
    answer: str
    original_answer: Optional[str] = None
    corrected_answer: Optional[str] = None
    flag_type: str = "untouched"
    confidence: Optional[float] = None
    availability: Optional[str] = None


class SubmissionIn(BaseModel):
    doc_id: str
    sheet_name: str
    items: list[ReviewItemIn]
    previous_submission_id: Optional[str] = None
    # Phase 5: header text of the mapped columns in the original workbook.
    # Optional so pre-Phase-5 clients still work; export requires them.
    question_col_name: Optional[str] = None
    availability_col_name: Optional[str] = None
    remarks_col_name: Optional[str] = None
    display_name: Optional[str] = None


class ItemDecision(BaseModel):
    question_id: str
    decision: str
    corrected_answer: Optional[str] = None
    comment: Optional[str] = None


class ApproveIn(BaseModel):
    # Per-item edits the reviewer made in the Review Queue. Keyed by
    # question_id; any qid not present is approved as-submitted. Missing
    # entirely = plain approve, no edits (matches historical behaviour).
    edits: dict[str, str] = Field(default_factory=dict)


class SendBackIn(BaseModel):
    decisions: list[ItemDecision]
    reviewer_comment: Optional[str] = None


def _notify(session: Session, user_email: str, ntype: str, message: str, link: str):
    session.add(Notification(user_email=user_email, type=ntype, message=message, link=link))


def _serialize(sub: ReviewSubmission, items: list[ReviewItem]) -> dict:
    return {
        "id": sub.id,
        "doc_id": sub.doc_id,
        "sheet_name": sub.sheet_name,
        "submitted_by": sub.submitted_by,
        "submitted_at": sub.submitted_at.isoformat() if sub.submitted_at else None,
        "status": sub.status,
        "reviewed_by": sub.reviewed_by,
        "reviewed_at": sub.reviewed_at.isoformat() if sub.reviewed_at else None,
        "reviewer_comment": sub.reviewer_comment,
        "previous_submission_id": sub.previous_submission_id,
        "cycle": sub.cycle,
        "question_col_name": sub.question_col_name,
        "availability_col_name": sub.availability_col_name,
        "remarks_col_name": sub.remarks_col_name,
        "display_name": sub.display_name,
        "items": [{
            "question_id": i.question_id,
            "question": i.question,
            "section": i.section,
            "answer": i.answer,
            "original_answer": i.original_answer,
            "corrected_answer": i.corrected_answer,
            "flag_type": i.flag_type,
            "decision": i.decision,
            "comment": i.comment,
            "confidence": i.confidence,
            "availability": i.availability,
        } for i in items],
        "counts": {
            "total": len(items),
            "corrected": sum(1 for i in items if i.flag_type == "corrected"),
            "flagged": sum(1 for i in items if i.flag_type == "flagged"),
            "accepted": sum(1 for i in items if i.flag_type == "accepted"),
        },
    }


@router.post("/submissions")
def create_submission(payload: SubmissionIn, user: User = Depends(require("generate"))):
    touched = [it for it in payload.items if it.flag_type in ("corrected", "flagged")]
    if not touched:
        raise HTTPException(400, "A submission needs at least one corrected or flagged answer")
    with Session(engine) as s:
        cycle = 1
        prev: Optional[ReviewSubmission] = None
        prev_id = payload.previous_submission_id
        if prev_id:
            prev = s.get(ReviewSubmission, prev_id)
            if prev:
                cycle = (prev.cycle or 1) + 1
        # Phase 5: on a resubmission (cycle > 1), the resubmit UI may not
        # re-send the column mapping — it's the same file. Inherit any
        # missing col_names from the previous submission so the export
        # endpoint has what it needs on the eventual approval.
        sub = ReviewSubmission(
            doc_id=payload.doc_id, sheet_name=payload.sheet_name,
            submitted_by=user.email, status="pending",
            previous_submission_id=prev_id, cycle=cycle,
            question_col_name=payload.question_col_name or (prev.question_col_name if prev else None),
            availability_col_name=payload.availability_col_name or (prev.availability_col_name if prev else None),
            remarks_col_name=payload.remarks_col_name or (prev.remarks_col_name if prev else None),
            display_name=payload.display_name or (prev.display_name if prev else None),
        )
        s.add(sub)
        s.flush()
        for it in payload.items:
            s.add(ReviewItem(
                submission_id=sub.id, question_id=it.question_id, question=it.question,
                section=it.section, answer=it.answer,
                original_answer=it.original_answer or it.answer,
                corrected_answer=it.corrected_answer, flag_type=it.flag_type,
                confidence=it.confidence, availability=it.availability,
            ))
        from auth import reviewer_emails
        label = sub.display_name or payload.sheet_name
        for r_email in reviewer_emails():
            if r_email != user.email.lower():
                _notify(s, r_email, "submission_received",
                        f"{user.name or user.email} sent '{label}' for review",
                        f"/review-queue?submission={sub.id}")
        s.commit()
        s.refresh(sub)
        items = s.exec(select(ReviewItem).where(ReviewItem.submission_id == sub.id)).all()
        return _serialize(sub, items)


@router.get("/submissions")
def list_submissions(user: User = Depends(current_user)):
    with Session(engine) as s:
        q = select(ReviewSubmission).order_by(ReviewSubmission.submitted_at.desc())
        if user.role not in ("reviewer", "admin"):
            q = q.where(ReviewSubmission.submitted_by == user.email)
        subs = s.exec(q).all()
        if not subs:
            return {"submissions": []}
        # Fetch ALL items for these submissions in ONE query, then group in
        # memory — avoids the N+1 round-trips to the database.
        sub_ids = [sub.id for sub in subs]
        all_items = s.exec(
            select(ReviewItem).where(ReviewItem.submission_id.in_(sub_ids))
        ).all()
        items_by_sub: dict[str, list[ReviewItem]] = {}
        for it in all_items:
            items_by_sub.setdefault(it.submission_id, []).append(it)
        out = [_serialize(sub, items_by_sub.get(sub.id, [])) for sub in subs]
        return {"submissions": out}


@router.get("/submissions/{submission_id}")
def get_submission(submission_id: str, user: User = Depends(current_user)):
    with Session(engine) as s:
        sub = s.get(ReviewSubmission, submission_id)
        if not sub:
            raise HTTPException(404, "Submission not found")
        if user.role not in ("reviewer", "admin") and sub.submitted_by != user.email:
            raise HTTPException(403, "Not your submission")
        items = s.exec(select(ReviewItem).where(ReviewItem.submission_id == sub.id)).all()
        return _serialize(sub, items)


@router.post("/submissions/{submission_id}/approve")
def approve_submission(
    submission_id: str,
    payload: Optional[ApproveIn] = Body(default=None),
    user: User = Depends(require("approve")),
):
    with Session(engine) as s:
        sub = s.get(ReviewSubmission, submission_id)
        if not sub:
            raise HTTPException(404, "Submission not found")
        if sub.status != "pending":
            raise HTTPException(400, f"Submission is already {sub.status}")
        items = s.exec(select(ReviewItem).where(ReviewItem.submission_id == sub.id)).all()

        # Apply reviewer edits BEFORE the ingest loop. An item the submitter
        # marked "accepted" but that the reviewer improved becomes a
        # correction — its flag_type gets promoted so the existing ingest
        # gate picks it up, and the FeedbackPair write from patch 17
        # naturally reflects the reviewer's improvement as good_answer.
        if payload and payload.edits:
            for it in items:
                edit = (payload.edits.get(it.question_id) or "").strip()
                if not edit:
                    continue
                original = (it.original_answer or it.answer or "").strip()
                if edit == original:
                    continue  # no-op edit, don't churn state
                it.corrected_answer = edit
                if it.flag_type == "accepted":
                    it.flag_type = "corrected"

        ingested = 0
        failed_ids: list[str] = []
        for it in items:
            answer_to_ingest = it.corrected_answer or (it.answer if it.flag_type == "corrected" else None)
            if it.flag_type in ("corrected", "flagged") and answer_to_ingest:
                try:
                    from ingest import ingest_correction
                    ingest_correction(it.question, answer_to_ingest, it.section or "", "review_queue")
                    it.decision = "approved"
                    ingested += 1
                    # Mirror the FeedbackPair write from /feedback/ingest so
                    # this approved correction appears in the Feedback Loop
                    # page (which reads feedback_pairs, not Qdrant). Same
                    # gate as ingest_correction — if the KB got it, the
                    # ledger records it too. Attribute to the submitter,
                    # stamp the approving reviewer.
                    s.add(FeedbackPair(
                        signal="correction",
                        question=it.question,
                        good_answer=answer_to_ingest,
                        # bad_answer must be the AI's first draft, not the
                        # submitter's already-corrected text. it.answer is
                        # `editedRemarks ?? remarks` from the frontend, so
                        # it already reflects submitter edits; the pristine
                        # first draft lives in it.original_answer.
                        bad_answer=it.original_answer or it.answer or None,
                        section=it.section or None,
                        source="review_queue",
                        user_email=sub.submitted_by,
                        user_name=sub.submitted_by,
                        reviewer_email=user.email,
                        reviewer_name=user.name or user.email,
                        confidence=float(it.confidence) if it.confidence is not None else None,
                    ))
                except Exception as e:
                    print(f"  [review] KB ingest failed for {it.question_id}: {e}")
                    failed_ids.append(it.question_id)
        # Data-integrity gate: if any intended KB ingest failed, do NOT
        # mark the sheet approved. Roll back so nothing (status, decisions,
        # FeedbackPairs) is committed, and tell the reviewer to retry. This
        # keeps approval and KB ingestion effectively atomic under Model A.
        if failed_ids:
            s.rollback()
            raise HTTPException(
                502,
                detail={
                    "message": "Knowledge-base ingestion failed for one or more "
                               "answers; the sheet was NOT approved. Please retry.",
                    "failed_question_ids": failed_ids,
                },
            )
        sub.status = "approved"
        sub.reviewed_by = user.email
        sub.reviewed_at = _now()
        # Don't self-notify: an admin who approves their own submission
        # shouldn't get an "approved" notification for it.
        if sub.submitted_by != user.email:
            _notify(s, sub.submitted_by, "submission_approved",
                    f"Your sheet '{sub.display_name or sub.sheet_name}' was approved",
                    f"/my-submissions?submission={sub.id}")
        s.commit()
        s.refresh(sub)
        return {"status": "approved", "ingested": ingested, "submission_id": sub.id}


@router.post("/submissions/{submission_id}/send-back")
def send_back_submission(submission_id: str, payload: SendBackIn, user: User = Depends(require("approve"))):
    rejected = [d for d in payload.decisions if d.decision == "rejected"]
    if not rejected:
        raise HTTPException(400, "Send-back requires at least one rejected answer")
    for d in rejected:
        if not (d.corrected_answer or d.comment):
            raise HTTPException(400, f"Rejected answer {d.question_id} needs a correction or comment")
    with Session(engine) as s:
        sub = s.get(ReviewSubmission, submission_id)
        if not sub:
            raise HTTPException(404, "Submission not found")
        if sub.status != "pending":
            raise HTTPException(400, f"Submission is already {sub.status}")
        items = {i.question_id: i for i in s.exec(select(ReviewItem).where(ReviewItem.submission_id == sub.id)).all()}
        for d in payload.decisions:
            it = items.get(d.question_id)
            if not it:
                continue
            it.decision = d.decision
            if d.corrected_answer:
                it.corrected_answer = d.corrected_answer
            if d.comment:
                it.comment = d.comment
        sub.status = "sent_back"
        sub.reviewed_by = user.email
        sub.reviewed_at = _now()
        sub.reviewer_comment = payload.reviewer_comment
        _notify(s, sub.submitted_by, "submission_sent_back",
                f"Your sheet '{sub.display_name or sub.sheet_name}' was sent back with feedback",
                f"/my-submissions?submission={sub.id}")
        s.commit()
        return {"status": "sent_back", "submission_id": sub.id}


def _safe_filename_stem(name: str) -> str:
    """Sanitize a user-provided string for use as a filename stem.

    Keeps letters, digits, spaces, hyphens, underscores, dots. Replaces
    everything else with underscore. Collapses whitespace runs. Strips
    leading/trailing whitespace and dots (some filesystems reject trailing
    dots). Returns 'export' if the result is empty.
    """
    cleaned = re.sub(r"[^A-Za-z0-9 \-_.]", "_", name)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    return cleaned or "export"


def _normalise_q(text: str) -> str:
    """Match the frontend exporter.ts normalise(): lowercase, collapse whitespace,
    trim, truncate to 200 chars. Same on both sides so lookups agree."""
    return re.sub(r"\s+", " ", (text or "").lower().strip())[:200]


@router.get("/submissions/{submission_id}/export")
def export_submission(submission_id: str, user: User = Depends(current_user)):
    """
    Phase 5: reopen the persisted original workbook and write approved
    availability + remarks into the mapped columns, preserving all other
    cells, formulas, formatting, and untouched sheets.

    Row-matching mirrors the frontend exporter.ts: normalise question text
    (lowercase, collapse whitespace, truncate to 200) and look up against
    ReviewItem.question. First cell match wins.

    Access: the submitter of the submission, or any reviewer/admin.
    Only approved submissions are exportable.
    """
    with Session(engine) as s:
        sub = s.get(ReviewSubmission, submission_id)
        if not sub:
            raise HTTPException(404, "Submission not found")
        if sub.submitted_by != user.email and user.role not in ("reviewer", "admin"):
            raise HTTPException(403, "Not your submission")
        if sub.status != "approved":
            raise HTTPException(400, f"Only approved submissions can be exported (this one is {sub.status})")
        if not (sub.question_col_name and sub.availability_col_name and sub.remarks_col_name):
            raise HTTPException(
                409,
                "This submission was created before write-back mapping was captured. "
                "Please re-upload the sheet, map your columns, and resubmit for review.",
            )
        original = s.get(OriginalDocument, sub.doc_id)
        if not original:
            raise HTTPException(
                409,
                "Original file no longer available. Please re-upload the sheet and resubmit.",
            )
        items = s.exec(select(ReviewItem).where(ReviewItem.submission_id == sub.id)).all()

    # openpyxl is imported lazily so unrelated endpoints don't pay the cost.
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(original.content))
    if sub.sheet_name not in wb.sheetnames:
        raise HTTPException(409, f"Sheet '{sub.sheet_name}' not found in the original workbook.")
    ws = wb[sub.sheet_name]

    # Locate the three mapped columns by exact header text (case-sensitive to
    # avoid false positives; the frontend captures the header text verbatim).
    q_col = a_col = r_col = None
    header_row = None
    for row_idx in range(1, min(9, (ws.max_row or 0) + 1)):
        for col_idx in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            txt = str(v).strip()
            if q_col is None and txt == sub.question_col_name:
                q_col = col_idx
                header_row = row_idx
            if a_col is None and txt == sub.availability_col_name:
                a_col = col_idx
                header_row = row_idx
            if r_col is None and txt == sub.remarks_col_name:
                r_col = col_idx
                header_row = row_idx
        if q_col and a_col and r_col:
            break

    missing = []
    if not q_col: missing.append(sub.question_col_name)
    if not a_col: missing.append(sub.availability_col_name)
    if not r_col: missing.append(sub.remarks_col_name)
    if missing:
        raise HTTPException(
            409,
            f"Column headers not found in the original sheet: {missing}. "
            "The file may have been modified since submission.",
        )

    # Build collision-safe lookups. Answers are resolved per question_id
    # (unique) — never by text alone — so duplicate/near-duplicate questions
    # can't clobber one another.
    #   * by_row: source_row -> (question_id, normalised question text). The
    #     source_row is the 1-based openpyxl row the parser captured, unique
    #     per item; the stored text lets us verify the row hasn't shifted.
    #   * text_queue: normalised text -> queue of question_ids, the fallback
    #     when a row shifted or an item has no source_row. Consuming from the
    #     queue means duplicate questions each claim a distinct row.
    from collections import defaultdict, deque

    ans_by_id: dict[str, tuple[str, str]] = {}
    by_row: dict[int, tuple[str, str]] = {}
    text_queue: dict[str, "deque[str]"] = defaultdict(deque)
    for it in items:
        avail = it.availability or ""
        remarks = it.corrected_answer or it.answer or ""
        norm = _normalise_q(it.question)
        ans_by_id[it.question_id] = (avail, remarks)
        if it.source_row:
            by_row[it.source_row] = (it.question_id, norm)
        text_queue[norm].append(it.question_id)

    # Walk data rows and write cells. Two robustness concerns:
    #   1. Cells inside a merged range are MergedCells — writing raises
    #      AttributeError. A merge across the answer columns typically
    #      means the row is a section-header banner (no answer belongs
    #      there), so we skip it cleanly rather than 500ing the export.
    #   2. Track which submission items didn't match any row so the caller
    #      can surface silent drops (returned in X-Items-Skipped header).
    from openpyxl.cell.cell import MergedCell

    injected = 0
    merged_skips = 0
    written_ids: set[str] = set()
    start = (header_row or 1) + 1
    for row_idx in range(start, (ws.max_row or 0) + 1):
        v = ws.cell(row=row_idx, column=q_col).value
        if v is None:
            continue
        key = _normalise_q(str(v))

        # Which item belongs on THIS row?
        #   1. Positional: the item whose source_row == this row, if its stored
        #      text still agrees (row hasn't shifted) and it's unclaimed.
        #   2. Fallback: the next unclaimed item queued under this row's text.
        qid = None
        cand = by_row.get(row_idx)
        if cand and cand[1] == key and cand[0] not in written_ids:
            qid = cand[0]
        if qid is None:
            queue = text_queue.get(key)
            while queue:
                nxt = queue.popleft()
                if nxt not in written_ids:
                    qid = nxt
                    break
        if qid is None:
            continue

        # If either target cell is inside a merged range, the row is a
        # section-header banner in the template — skip rather than trying
        # to unmerge (that would clobber the customer's layout).
        a_cell = ws.cell(row=row_idx, column=a_col)
        r_cell = ws.cell(row=row_idx, column=r_col)
        if isinstance(a_cell, MergedCell) or isinstance(r_cell, MergedCell):
            merged_skips += 1
            continue

        avail, remarks = ans_by_id[qid]
        a_cell.value = avail
        r_cell.value = remarks
        injected += 1
        written_ids.add(qid)

    # Items never written to a row: question text drifted / row removed, or
    # they landed on a merged section-header row we skipped above.
    skipped_ids = [it.question_id for it in items if it.question_id not in written_ids]

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    # Prefer the user-provided display name for a nicer download filename.
    # Fall back to the uploaded file's own name for legacy submissions.
    stem = _safe_filename_stem(sub.display_name) if sub.display_name else (Path(original.filename).stem or "export")
    filename = f"{stem}_answered.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Items-Total": str(len(items)),
            "X-Items-Written": str(injected),
            "X-Items-Skipped": ",".join(skipped_ids) if skipped_ids else "",
            "X-Rows-Skipped-Merged": str(merged_skips),
        },
    )


# ── Notifications (in-app) ────────────────────────────────────────────────────

@router.get("/notifications")
def get_notifications(user: User = Depends(current_user)):
    with Session(engine) as s:
        rows = s.exec(
            select(Notification)
            .where(Notification.user_email == user.email)
            .order_by(Notification.created_at.desc())
        ).all()
        return {
            "notifications": [{
                "id": n.id, "type": n.type, "message": n.message,
                "link": n.link, "read": n.read,
                "created_at": n.created_at.isoformat() if n.created_at else None,
            } for n in rows],
            "unread": sum(1 for n in rows if not n.read),
        }


@router.post("/notifications/{notification_id}/read")
def mark_notification_read(notification_id: str, user: User = Depends(current_user)):
    with Session(engine) as s:
        n = s.get(Notification, notification_id)
        if not n or n.user_email != user.email:
            raise HTTPException(404, "Notification not found")
        n.read = True
        s.commit()
        return {"ok": True}


@router.post("/notifications/read-all")
def mark_all_read(user: User = Depends(current_user)):
    with Session(engine) as s:
        rows = s.exec(
            select(Notification).where(
                Notification.user_email == user.email, Notification.read == False  # noqa: E712
            )
        ).all()
        for n in rows:
            n.read = True
        s.commit()
        return {"marked": len(rows)}


# ── Dashboard stats (computed server-side, single source of truth) ────────────

# Time-saved assumptions — EXPLICIT and adjustable. These are estimates, not
# measured facts; keep them conservative and defensible.
MANUAL_HOURS_PER_RFP = 16.0      # ~2 working days to fill an RFP by hand
GENERATION_HOURS_PER_RFP = 0.25  # ~15 min for AI generation (tool cost, ex-review)
WORKDAY_HOURS = 8.0


def _median(xs: list[float]) -> float:
    if not xs:
        return 0.0
    s = sorted(xs)
    n = len(s)
    mid = n // 2
    return s[mid] if n % 2 else (s[mid - 1] + s[mid]) / 2


@router.get("/dashboard-stats")
def dashboard_stats(user: User = Depends(current_user)):
    with Session(engine) as s:
        subs = {x.id: x for x in s.exec(select(ReviewSubmission)).all()}
        all_subs = list(subs.values())

        # RFPs processed = every uploaded document (each upload event counts),
        # including ones that were generated but never sent for review.
        rfps_processed = len(s.exec(select(OriginalDocument)).all())

        # A document is "in review" if the latest cycle in its lineage is pending.
        # A document is "completed" if any cycle reached approved.
        # Group submissions by doc_id, pick lineage state.
        by_doc: dict[str, list[ReviewSubmission]] = {}
        for x in all_subs:
            by_doc.setdefault(x.doc_id, []).append(x)

        in_review_docs = 0
        completed_docs = 0
        for doc_id, group in by_doc.items():
            statuses = {g.status for g in group}
            if "approved" in statuses:
                completed_docs += 1
            elif "pending" in statuses:
                in_review_docs += 1

        # Review turnaround: for each APPROVED submission, walk back to the first
        # cycle in its lineage; turnaround = approved.reviewed_at - first.submitted_at.
        # Use MEDIAN (robust to sheets left sitting for days).
        turnarounds_min: list[float] = []
        for a in all_subs:
            if a.status != "approved" or not a.reviewed_at:
                continue
            first = a
            seen = set()
            while first.previous_submission_id and first.previous_submission_id in subs \
                    and first.id not in seen:
                seen.add(first.id)
                first = subs[first.previous_submission_id]
            if first.submitted_at:
                delta_min = (a.reviewed_at - first.submitted_at).total_seconds() / 60.0
                if delta_min >= 0:
                    turnarounds_min.append(delta_min)
        median_turnaround_min = _median(turnarounds_min)

        # Time saved = completed RFPs * (manual - generation), in hours -> days.
        hours_saved = completed_docs * (MANUAL_HOURS_PER_RFP - GENERATION_HOURS_PER_RFP)
        days_saved = hours_saved / WORKDAY_HOURS

        # Avg confidence across all items.
        items = s.exec(select(ReviewItem)).all()
        confs = [i.confidence for i in items if i.confidence and i.confidence > 0]
        avg_confidence = sum(confs) / len(confs) if confs else 0.0

        return {
            "rfps_processed": rfps_processed,
            "in_review": in_review_docs,
            "completed": completed_docs,
            "days_saved": round(days_saved, 1),
            "hours_saved": round(hours_saved, 1),
            "median_review_minutes": round(median_turnaround_min, 1),
            "avg_confidence": round(avg_confidence, 3),
            "assumptions": {
                "manual_hours_per_rfp": MANUAL_HOURS_PER_RFP,
                "generation_hours_per_rfp": GENERATION_HOURS_PER_RFP,
            },
        }
